import os
import datetime as dt
from openpyxl import load_workbook
import pandas as pd
import numpy as np

base = r"c:\Users\PC\Desktop\JANUARY TAT 2026"
path_daily = os.path.join(base,"JAN DAILY TAT.xlsx")
path_completed = os.path.join(base,"JAN COMPLETED TOKENS TAT 2026.xlsx")

cols_daily = ["Date","Month","Token","Modality","Issued At","Status","Turnaround Time","TARGET TAT","Wait Time + Service time"]
cols_comp = ["Date","Month","Token","Modality","Issued At","Status","ACTUAL Turnaround Time","TARGET TAT","SERVICE + WAIT TIME"]

start = dt.datetime(2026,1,1)
end = dt.datetime(2026,2,1)

def read_filtered(path, wanted_cols):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    use_idx = [idx[c] for c in wanted_cols if c in idx]
    use_cols = [c for c in wanted_cols if c in idx]
    if not use_idx:
        wb.close()
        return pd.DataFrame(columns=use_cols)

    max_col = max(use_idx) + 1  # openpyxl is 1-based for max_col
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=max_col):
        date_cell = row[idx['Date']].value if 'Date' in idx else None
        if date_cell is None:
            continue
        if isinstance(date_cell, dt.datetime):
            date_val = date_cell
        elif isinstance(date_cell, dt.date):
            date_val = dt.datetime.combine(date_cell, dt.time())
        else:
            try:
                date_val = pd.to_datetime(date_cell)
            except Exception:
                continue
        if not (start <= date_val < end):
            continue
        vals = [row[i].value for i in use_idx]
        rows.append(vals)
    wb.close()
    return pd.DataFrame(rows, columns=use_cols)

jan_daily = read_filtered(path_daily, cols_daily)
jan_comp = read_filtered(path_completed, cols_comp)

# helper to convert excel time or string to minutes

def to_minutes(series):
    s = series.copy()
    if s is None:
        return pd.Series(dtype=float)
    if pd.api.types.is_timedelta64_dtype(s):
        return s.dt.total_seconds()/60
    s_num = pd.to_numeric(s, errors='coerce')
    minutes_from_num = s_num * 24 * 60
    s_str = s.astype(str)
    td = pd.to_timedelta(s_str, errors='coerce')
    minutes_from_str = td.dt.total_seconds()/60
    return minutes_from_str.fillna(minutes_from_num)

jan_comp['actual_min'] = to_minutes(jan_comp.get('ACTUAL Turnaround Time'))
jan_comp['target_min'] = to_minutes(jan_comp.get('TARGET TAT'))
jan_comp['svc_wait_min'] = to_minutes(jan_comp.get('SERVICE + WAIT TIME'))

jan_daily['tat_min'] = to_minutes(jan_daily.get('Turnaround Time'))
jan_daily['target_min'] = to_minutes(jan_daily.get('TARGET TAT'))
jan_daily['svc_wait_min'] = to_minutes(jan_daily.get('Wait Time + Service time'))

# status mix
status_mix = jan_daily['Status'].value_counts(dropna=False).rename('count').to_frame()
status_mix['pct'] = (status_mix['count'] / status_mix['count'].sum() * 100).round(1)

# completion rate
completed_count = (jan_daily['Status'].astype(str).str.lower() == 'complete').sum()
completion_rate = completed_count / len(jan_daily) if len(jan_daily) else np.nan

# overall TAT stats for completed tokens
jan_comp['within_target'] = (jan_comp['actual_min'] <= jan_comp['target_min'])

summary = {
    'jan_daily_rows': len(jan_daily),
    'jan_completed_rows': len(jan_comp),
    'completion_rate': completion_rate,
    'overall_median_actual_min': jan_comp['actual_min'].median(),
    'overall_mean_actual_min': jan_comp['actual_min'].mean(),
    'overall_p90_actual_min': jan_comp['actual_min'].quantile(0.9),
    'overall_compliance': jan_comp['within_target'].mean(),
}

modality_summary = jan_comp.groupby('Modality').agg(
    tokens=('Token','count'),
    median_actual_min=('actual_min','median'),
    mean_actual_min=('actual_min','mean'),
    p90_actual_min=('actual_min', lambda x: x.quantile(0.9)),
    target_min=('target_min','median'),
    compliance=('within_target','mean')
).sort_values('tokens', ascending=False)

trend = jan_comp.groupby('Date').agg(
    tokens=('Token','count'),
    median_actual_min=('actual_min','median'),
    mean_actual_min=('actual_min','mean'),
    compliance=('within_target','mean')
).reset_index().sort_values('Date')

# Save outputs
out_dir = base
status_mix.to_csv(os.path.join(out_dir, "status_mix.csv"))
modality_summary.to_csv(os.path.join(out_dir, "modality_summary.csv"))
trend.to_csv(os.path.join(out_dir, "daily_trend.csv"), index=False)

with open(os.path.join(out_dir, "summary.txt"), "w", encoding="utf-8") as f:
    for k,v in summary.items():
        f.write(f"{k}: {v}\n")
