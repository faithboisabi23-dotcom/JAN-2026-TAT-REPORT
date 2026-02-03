import os
import datetime as dt
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt

base = r"c:\Users\PC\Desktop\JANUARY TAT 2026"
path_daily = os.path.join(base, "JAN DAILY TAT.xlsx")

start = dt.datetime(2026, 1, 1)
end = dt.datetime(2026, 2, 1)

wanted_cols = ["Date", "Token", "Modality", "Status"]

wb = load_workbook(path_daily, read_only=True, data_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {h: i for i, h in enumerate(headers)}
use_idx = [idx[c] for c in wanted_cols if c in idx]
use_cols = [c for c in wanted_cols if c in idx]
if not use_idx:
    wb.close()
    raise RuntimeError("Required columns not found.")

max_col = max(use_idx) + 1
rows = []
for row in ws.iter_rows(min_row=2, max_col=max_col):
    date_cell = row[idx['Date']].value if 'Date' in idx else None
    if date_cell is None:
        continue
    if isinstance(date_cell, dt.datetime):
        date_val = date_cell
    elif isinstance(date_cell, dt.date):
        date_val = dt.datetime.combine(date_cell, dt.time())
    else:
        try:
            date_val = pd.to_datetime(date_cell)
        except Exception:
            continue
    if not (start <= date_val < end):
        continue
    vals = [row[i].value for i in use_idx]
    rows.append(vals)

wb.close()

jan_daily = pd.DataFrame(rows, columns=use_cols)

# Normalize modality codes and status labels
jan_daily['Modality'] = jan_daily['Modality'].astype(str).str.strip().str.upper()
status = jan_daily['Status'].astype(str).str.strip().str.title()
status = status.replace({"E. Complete": "E-Complete", "NoShow": "Noshow", "No Show": "Noshow"})
jan_daily['Status'] = status

# Overall status counts
status_counts = jan_daily['Status'].value_counts().rename('count')
completion_rate = status_counts.get('Complete', 0) / len(jan_daily) if len(jan_daily) else np.nan

scorecard = pd.DataFrame({
    'metric': [
        'Total tokens',
        'Completed tokens',
        'Pending tokens',
        'Serving tokens',
        'E-Complete tokens',
        'No-show tokens',
        'Standby tokens',
        'Completion rate'
    ],
    'value': [
        len(jan_daily),
        int(status_counts.get('Complete', 0)),
        int(status_counts.get('Pending', 0)),
        int(status_counts.get('Serving', 0)),
        int(status_counts.get('E-Complete', 0)),
        int(status_counts.get('Noshow', 0)),
        int(status_counts.get('Standby', 0)),
        completion_rate
    ]
})

# Modality-specific counts
focus = {'XR': 'XRAY', 'MR': 'MRI', 'CT': 'CT', 'US': 'Ultrasound'}
modality_rows = []
modality_detail_rows = []
for code, label in focus.items():
    df_m = jan_daily[jan_daily['Modality'] == code]
    total = len(df_m)
    completed_m = (df_m['Status'] == 'Complete').sum()
    non_completed_m = df_m['Status'].isin(['Pending', 'Serving', 'Noshow', 'Standby']).sum()
    pending_m = (df_m['Status'] == 'Pending').sum()
    serving_m = (df_m['Status'] == 'Serving').sum()
    noshow_m = (df_m['Status'] == 'Noshow').sum()
    standby_m = (df_m['Status'] == 'Standby').sum()
    ecomplete_m = (df_m['Status'] == 'E-Complete').sum()
    modality_rows.append({
        'Modality': label,
        'Total tokens': total,
        'Completed tokens': int(completed_m),
        'Non-completed tokens': int(non_completed_m)
    })
    modality_detail_rows.append({
        'Modality': label,
        'Total tokens': total,
        'Completed': int(completed_m),
        'Pending': int(pending_m),
        'Serving': int(serving_m),
        'Noshow': int(noshow_m),
        'Standby': int(standby_m),
        'E-Complete': int(ecomplete_m)
    })

modality_summary = pd.DataFrame(modality_rows)
modality_detail = pd.DataFrame(modality_detail_rows)

# Save outputs
out_dir = base
scorecard.to_csv(os.path.join(out_dir, "monthly_token_scorecard.csv"), index=False)
modality_summary.to_csv(os.path.join(out_dir, "modality_token_status.csv"), index=False)
modality_detail.to_csv(os.path.join(out_dir, "modality_token_status_detail.csv"), index=False)

# Save charts
chart_dir = os.path.join(base, "charts")
os.makedirs(chart_dir, exist_ok=True)
for _, row in modality_summary.iterrows():
    label = row['Modality']
    completed = row['Completed tokens']
    non_completed = row['Non-completed tokens']
    plt.figure(figsize=(4, 3))
    plt.bar(['Completed', 'Non-completed'], [completed, non_completed], color=['#2ca02c', '#d62728'])
    plt.title(f"{label} Tokens: Completed vs Non-completed")
    plt.ylabel("Tokens")
    plt.tight_layout()
    out_path = os.path.join(chart_dir, f"{label.lower().replace(' ', '_')}_status.png")
    plt.savefig(out_path, dpi=200)
    plt.close()
