import os
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from urllib.parse import quote
import datetime as dt
from openpyxl import load_workbook

base = r"c:\Users\PC\Desktop\JANUARY TAT 2026"
logo_file = "SONAR LOGO 2024 (1)1 (1).png"
logo_src = quote(logo_file)
daily_path = os.path.join(base, "JAN DAILY TAT.xlsx")
completed_path = os.path.join(base, "JAN COMPLETED TOKENS TAT 2026.xlsx")
scorecard_path = os.path.join(base, "monthly_token_scorecard.csv")
modality_status_path = os.path.join(base, "modality_token_status.csv")
modality_detail_path = os.path.join(base, "modality_token_status_detail.csv")
trend_path = os.path.join(base, "daily_trend.csv")
modality_tat_path = os.path.join(base, "modality_summary.csv")

scorecard = pd.read_csv(scorecard_path)
mod_status = pd.read_csv(modality_status_path)
mod_detail = pd.read_csv(modality_detail_path)
mod_tat = pd.read_csv(modality_tat_path)

start = dt.datetime(2026, 1, 1)
end = dt.datetime(2026, 2, 1)


def read_filtered_xlsx(path, wanted_cols):
  wb = load_workbook(path, read_only=True, data_only=True)
  ws = wb.active
  rows = ws.iter_rows(values_only=True)
  headers = next(rows, None)
  if not headers:
    hours = total // 60
    mins = total % 60
    return f"{hours:02d}:{mins:02d}"
        <li>MRI has the largest incomplete workload; XRAY shows moderate incomplete load.</li>
        <li>Prioritize MRI throughput to stabilize overall completion.</li>
      </ul>
    </div>

    <div class="card">
      <div id="tat_vs_target"></div>
      <ul class="bullets">
        <li>All four modalities exceed target TAT on average.</li>
        <li>MRI shows the widest gap, indicating a capacity constraint.</li>
      </ul>
    </div>

    <div class="card">
      <div id="combo_xray"></div>
      <ul class="bullets">
        <li>Daily XRAY completed volume vs average TAT with target reference.</li>
      </ul>
    </div>

    <div class="card">
      <div id="combo_mri"></div>
      <ul class="bullets">
        <li>Daily MRI completed volume vs average TAT with target reference.</li>
      </ul>
    </div>

    <div class="card">
      <div id="combo_ct"></div>
      <ul class="bullets">
        <li>Daily CT completed volume vs average TAT with target reference.</li>
      </ul>
    </div>

    <div class="card">
      <div id="combo_us"></div>
      <ul class="bullets">
        <li>Daily Ultrasound completed volume vs average TAT with target reference.</li>
      </ul>
    </div>
      marker_color="#d9d9d9",
      text=counts,
      textposition="outside"
    ),
    secondary_y=False
  )
  fig.add_trace(
    go.Scatter(
      x=dates,
      y=avg_tat,
      name="Avg Actual TAT (min)",
      mode="lines+markers",
      line=dict(color="#d62728", width=2),
      connectgaps=True,
      hovertemplate="%{x|%b %d}: %{customdata}"
    ),
    secondary_y=True
  )
  fig.data[-1].customdata = [minutes_to_hhmm(v) for v in avg_tat]
  fig.add_trace(
    go.Scatter(
      x=dates,
      y=target_line,
      name="Target TAT (min)",
      mode="lines",
      line=dict(color="#2ca02c", width=2, dash="solid"),
      hovertemplate="%{x|%b %d}: %{customdata}"
    ),
    secondary_y=True
  )
  fig.data[-1].customdata = [minutes_to_hhmm(v) for v in target_line]

  fig.update_layout(
    title=f"{modality_label} TAT Trend",
    height=360,
    margin=dict(l=20, r=20, t=30, b=40),
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0)
  )
  fig.update_yaxes(title_text="Completed tokens", secondary_y=False)
  max_tat = np.nanmax([*avg_tat, *target_line]) if len(avg_tat) else np.nan
  tickvals, ticktext = build_time_ticks(max_tat)
  fig.update_yaxes(title_text="TAT (HH:MM)", secondary_y=True, tickvals=tickvals, ticktext=ticktext)
  return fig

# Scorecards
scorecard_vals = scorecard.copy()
scorecard_vals["value"] = scorecard_vals["value"].astype(float)

# Build scorecard figure (2 rows x 4 columns)
score_fig = make_subplots(rows=2, cols=4, specs=[[{"type": "indicator"}]*4]*2)

labels = scorecard_vals["metric"].tolist()
values = scorecard_vals["value"].tolist()

for i, (label, value) in enumerate(zip(labels, values)):
    r = i // 4 + 1
    c = i % 4 + 1
    display_value = value
    if label == "Completion rate":
        display_value = value * 100
        number = {"suffix": "%", "valueformat": ".1f"}
    else:
        number = {"valueformat": ",.0f"}
    score_fig.add_trace(
        go.Indicator(mode="number", value=display_value, title={"text": label}, number=number),
        row=r, col=c
    )

score_fig.update_layout(height=350, margin=dict(l=20, r=20, t=20, b=10))

# Modality status chart (Completed vs Non-completed)
mod_status_fig = go.Figure()
mod_status_fig.add_trace(go.Bar(
    x=mod_status["Modality"], y=mod_status["Completed tokens"], name="Completed", marker_color="#2ca02c"
))
mod_status_fig.add_trace(go.Bar(
    x=mod_status["Modality"], y=mod_status["Non-completed tokens"], name="Non-completed", marker_color="#d62728"
))
mod_status_fig.update_layout(
    barmode="group",
    height=360,
    margin=dict(l=20, r=20, t=30, b=20),
    title="Completed vs Non-completed Tokens by Modality"
)

# Modality TAT vs Target
focus = ["US", "MR", "XR", "CT"]
mod_tat_focus = mod_tat[mod_tat["Modality"].isin(focus)].copy()
mod_map = {"US": "Ultrasound", "MR": "MRI", "XR": "XRAY", "CT": "CT"}
mod_tat_focus["Modality"] = mod_tat_focus["Modality"].map(mod_map)

mod_tat_fig = go.Figure()
mod_tat_fig.add_trace(go.Bar(
  x=mod_tat_focus["Modality"], y=mod_tat_focus["mean_actual_min"], name="Average TAT", marker_color="#636efa",
  text=[minutes_to_hhmm(v) for v in mod_tat_focus["mean_actual_min"]],
  textposition="outside",
  hovertemplate="%{x}: %{customdata}"
))
mod_tat_fig.data[-1].customdata = [minutes_to_hhmm(v) for v in mod_tat_focus["mean_actual_min"]]
mod_tat_fig.add_trace(go.Bar(
  x=mod_tat_focus["Modality"], y=mod_tat_focus["target_min"], name="Target TAT", marker_color="#00cc96",
  text=[minutes_to_hhmm(v) for v in mod_tat_focus["target_min"]],
  textposition="outside",
  hovertemplate="%{x}: %{customdata}"
))
mod_tat_fig.data[-1].customdata = [minutes_to_hhmm(v) for v in mod_tat_focus["target_min"]]
mod_tat_fig.update_layout(
    barmode="group",
    height=360,
    margin=dict(l=20, r=20, t=30, b=20),
    title="Average TAT vs Target by Modality"
)
mod_max = np.nanmax([mod_tat_focus["mean_actual_min"].max(), mod_tat_focus["target_min"].max()])
mod_tickvals, mod_ticktext = build_time_ticks(mod_max)
mod_tat_fig.update_yaxes(title_text="TAT (HH:MM)", tickvals=mod_tickvals, ticktext=mod_ticktext)

# Daily combo charts by modality
combo_xray = build_combo_chart("XR", "XRAY")
combo_mri = build_combo_chart("MR", "MRI")
combo_ct = build_combo_chart("CT", "CT")
combo_us = build_combo_chart("US", "Ultrasound")

# Build HTML
html = f"""
<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"UTF-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\" />
  <title>January 2026 TAT & Token Workflow Report</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 0; background: #f7f7f9; color: #222; }}
    header {{ background: #0b3d91; color: #fff; padding: 16px 32px; display: flex; align-items: center; justify-content: space-between; gap: 16px; }}
    header img {{ height: 44px; width: auto; background: #fff; border-radius: 6px; padding: 4px; }}
    h1 {{ margin: 0; font-size: 22px; }}
    .container {{ padding: 20px 28px 40px; }}
    .card {{ background: #fff; border-radius: 8px; padding: 16px; margin-bottom: 18px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
    .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
    .bullets {{ margin: 8px 0 0 0; padding-left: 18px; }}
    .bullets li {{ margin: 4px 0; }}
    @media (max-width: 1100px) {{ .grid {{ grid-template-columns: 1fr; }} }}
  </style>
  <script src=\"https://cdn.plot.ly/plotly-2.32.0.min.js\"></script>
</head>
<body>
  <header>
    <h1>January 2026 TAT & Token Workflow Report</h1>
    <img src=\"{logo_src}\" alt=\"Sonar Imaging Centre\" />
  </header>
  <div class=\"container\">
    <div class="card">
      <div id="modality_status"></div>
      <ul class="bullets">
        <li>MRI has the largest incomplete workload; XRAY shows moderate incomplete load.</li>
        <li>Prioritize MRI throughput to stabilize overall completion.</li>
      </ul>
    </div>
      <div class=\"card\">
        <div id=\"status_mix\"></div>
        <ul class=\"bullets\">
          <li>Standby tokens are concentrated in MRI, indicating persistent queue pressure.</li>
          <li>No-shows are most visible in MRI and Ultrasound, increasing schedule volatility.</li>
        </ul>
      </div>
    </div>

    <div class=\"card\">
      <div id=\"tat_vs_target\"></div>
      <ul class=\"bullets\">
        <li>All four modalities exceed target TAT on average.</li>
        <li>MRI shows the widest gap, indicating a capacity constraint.</li>
      </ul>
    </div>

    <div class=\"card\">
      <div id=\"combo_xray\"></div>
      <ul class=\"bullets\">
        <li>Daily XRAY completed volume vs average TAT with target reference.</li>
      </ul>
    </div>

    <div class=\"card\">
      <div id=\"combo_mri\"></div>
      <ul class=\"bullets\">
        <li>Daily MRI completed volume vs average TAT with target reference.</li>
      </ul>
    </div>

    <div class=\"card\">
      <div id=\"combo_ct\"></div>
      <ul class=\"bullets\">
        <li>Daily CT completed volume vs average TAT with target reference.</li>
      </ul>
    </div>

    <div class=\"card\">
      <div id=\"combo_us\"></div>
      <ul class=\"bullets\">
        <li>Daily Ultrasound completed volume vs average TAT with target reference.</li>
      </ul>
    </div>
  </div>

  <script>
    var scoreData = {score_fig.to_json()};
    var modStatusData = {mod_status_fig.to_json()};
    var tatVsTargetData = {mod_tat_fig.to_json()};
    var comboXrayData = {combo_xray.to_json()};
    var comboMriData = {combo_mri.to_json()};
    var comboCtData = {combo_ct.to_json()};
    var comboUsData = {combo_us.to_json()};

    Plotly.newPlot('scorecards', scoreData.data, scoreData.layout, {{displayModeBar: false}});
    Plotly.newPlot('modality_status', modStatusData.data, modStatusData.layout, {{displayModeBar: false}});
    Plotly.newPlot('tat_vs_target', tatVsTargetData.data, tatVsTargetData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_xray', comboXrayData.data, comboXrayData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_mri', comboMriData.data, comboMriData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_ct', comboCtData.data, comboCtData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_us', comboUsData.data, comboUsData.layout, {{displayModeBar: false}});
  </script>
</body>
</html>
"""

with open(os.path.join(base, "dashboard.html"), "w", encoding="utf-8") as f:
    f.write(html)
