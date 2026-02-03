import os
import datetime as dt
from urllib.parse import quote

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook

base = r"c:\Users\PC\Desktop\JANUARY TAT 2026"
logo_file = "SONAR LOGO 2024 (1)1 (1).png"
logo_src = quote(logo_file)
completed_path = os.path.join(base, "JAN COMPLETED TOKENS TAT 2026.xlsx")
scorecard_path = os.path.join(base, "monthly_token_scorecard.csv")
modality_status_path = os.path.join(base, "modality_token_status.csv")
modality_tat_path = os.path.join(base, "modality_summary.csv")


def read_filtered_xlsx(path, wanted_cols):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        wb.close()
        return pd.DataFrame(columns=wanted_cols)

    col_idx = {h: i for i, h in enumerate(headers)}
    available = [c for c in wanted_cols if c in col_idx]
    data = {c: [] for c in available}

    for row in rows:
        for c in available:
            data[c].append(row[col_idx[c]])

    wb.close()
    return pd.DataFrame(data)


def to_minutes(value):
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    if isinstance(value, dt.time):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 60
    if isinstance(value, str):
        try:
            td = pd.to_timedelta(value)
            return td.total_seconds() / 60
        except Exception:
            return np.nan
    try:
        td = pd.to_timedelta(value)
        return td.total_seconds() / 60
    except Exception:
        return np.nan


def minutes_to_hhmm(minutes):
    if minutes is None or (isinstance(minutes, float) and np.isnan(minutes)):
        return ""
    total = int(round(float(minutes)))
    hours = total // 60
    mins = total % 60
    return f"{hours:02d}:{mins:02d}"


def build_time_ticks(max_minutes, step=60):
    if pd.isna(max_minutes) or max_minutes <= 0:
        return [], []
    max_val = int(np.ceil(max_minutes / step) * step)
    tickvals = list(range(0, max_val + step, step))
    ticktext = [minutes_to_hhmm(v) for v in tickvals]
    return tickvals, ticktext


def normalize_status(series):
    s = series.astype(str).str.strip().str.title()
    return s.replace({"E. Complete": "E-Complete", "NoShow": "Noshow", "No Show": "Noshow"})


STAGE_COLUMNS = {
    "XR": {
        "Billing": ["TAT - BILLING XRAY"],
        "Service": ["XRAY - TAT SERVICE"],
        "Dispatch": ["TAT - DISPATCH SERVICE"],
    },
    "MR": {
        "Billing": ["TAT - MRI BILLING"],
        "Service": ["TAT - MRI SEVICE"],
        "Dispatch": ["TAT - DISPATCH SERVICE"],
    },
    "CT": {
        "Billing": ["TAT - CT BILLING"],
        "Service": ["TAT - CT SERVICE"],
        "Dispatch": ["TAT - DISPATCH SERVICE"],
    },
    "US": {
        "Billing": ["TAT - U/S BILLING"],
        "Service": ["TAT - US SERVICE"],
        "Dispatch": ["TAT - DISPATCH SERVICE"],
    },
}


def average_stage_minutes(df, columns):
    values = []
    for col in columns:
        if col in df.columns:
            values.append(df[col].apply(to_minutes))
    if not values:
        return np.nan
    combined = pd.concat(values, axis=1)
    row_sums = combined.sum(axis=1, skipna=True)
    return row_sums.mean()


def compute_tat_distribution(modality_code):
    stage_cols = STAGE_COLUMNS[modality_code]
    wanted_cols = ["Date", "Modality", "Status"]
    for cols in stage_cols.values():
        wanted_cols.extend(cols)

    df = read_filtered_xlsx(completed_path, wanted_cols)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df[(df["Date"] >= dt.datetime(2026, 1, 1)) & (df["Date"] < dt.datetime(2026, 2, 1))]
    df["Modality"] = df["Modality"].astype(str).str.strip().str.upper()
    if "Status" in df.columns:
        df["Status"] = normalize_status(df["Status"])
        df = df[df["Status"].isin(["Complete", "E-Complete"])]
    df = df[df["Modality"] == modality_code]

    stage_minutes = {}
    for stage, cols in stage_cols.items():
        series_parts = []
        for col in cols:
            if col in df.columns:
                series_parts.append(df[col].apply(to_minutes))
        if series_parts:
            combined = pd.concat(series_parts, axis=1)
            stage_minutes[stage] = combined.sum(axis=1, skipna=True)
        else:
            stage_minutes[stage] = pd.Series([np.nan] * len(df), index=df.index)

    if stage_minutes:
        stage_frame = pd.DataFrame(stage_minutes)
        nonzero_mask = stage_frame.fillna(0).sum(axis=1) > 0
        df = df[nonzero_mask]
        stage_frame = stage_frame[nonzero_mask]
    else:
        stage_frame = pd.DataFrame(index=df.index)

    stage_avgs = {stage: (stage_frame[stage].mean() if stage in stage_frame else 0) for stage in stage_cols}
    stage_avgs = {k: (0 if pd.isna(v) else v) for k, v in stage_avgs.items()}

    total_all = np.nansum(list(stage_avgs.values()))
    if total_all <= 0:
        percentages = {stage: 0 for stage in stage_avgs}
    else:
        percentages = {stage: (val / total_all) * 100 for stage, val in stage_avgs.items()}
    return stage_avgs, percentages, len(df)


def build_combo_chart(
    modality_code,
    modality_label,
    height=300,
    legend_y=1.02,
    title_y=0.98,
    legend_x=0,
    legend_xanchor="left",
    legend_yanchor="top",
    margin_top=30,
    legend_font_size=11,
    bar_headroom=0.0,
    margin_right=10,
    x_pad_days=0,
    bar_cliponaxis=True,
    x_automargin=False,
):
    comp_df = read_filtered_xlsx(
        completed_path,
        ["Date", "Modality", "Status", "ACTUAL Turnaround Time", "TARGET TAT"],
    )
    comp_df["Date"] = pd.to_datetime(comp_df["Date"], errors="coerce")
    comp_df["Modality"] = comp_df["Modality"].astype(str).str.strip().str.upper()
    if "Status" in comp_df.columns:
        comp_df["Status"] = normalize_status(comp_df["Status"])
        comp_df = comp_df[comp_df["Status"].isin(["Complete", "E-Complete"])]

    comp_df = comp_df[comp_df["Modality"] == modality_code]
    comp_df["actual_min"] = comp_df["ACTUAL Turnaround Time"].apply(to_minutes)
    comp_df["target_min"] = comp_df["TARGET TAT"].apply(to_minutes)

    daily_counts = comp_df.groupby(comp_df["Date"].dt.date).size()
    daily_avg = comp_df.groupby(comp_df["Date"].dt.date)["actual_min"].mean()
    target_val = comp_df["target_min"].median()

    dates = [pd.Timestamp(d) for d in daily_counts.index]
    counts = daily_counts.values.tolist()
    avg_tat = [daily_avg.get(d, np.nan) for d in daily_counts.index]
    target_line = [target_val for _ in daily_counts.index]

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(
            x=dates,
            y=counts,
            name="Completed tokens",
            marker_color="#d9d9d9",
            text=counts,
            textposition="outside",
            cliponaxis=bar_cliponaxis,
        ),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(
            x=dates,
            y=avg_tat,
            name="Avg Actual TAT (min)",
            mode="lines+markers",
            line=dict(color="#d62728", width=2),
            connectgaps=True,
            hovertemplate="%{x|%b %d}: %{customdata}",
        ),
        secondary_y=True,
    )
    fig.data[-1].customdata = [minutes_to_hhmm(v) for v in avg_tat]

    fig.add_trace(
        go.Scatter(
            x=dates,
            y=target_line,
            name="Target TAT (min)",
            mode="lines",
            line=dict(color="#2ca02c", width=2, dash="solid"),
            hovertemplate="%{x|%b %d}: %{customdata}",
        ),
        secondary_y=True,
    )
    fig.data[-1].customdata = [minutes_to_hhmm(v) for v in target_line]

    max_tat = np.nanmax([*avg_tat, *target_line]) if len(avg_tat) else np.nan
    fig.update_layout(
        title=dict(text=f"{modality_label} TAT Trend", y=title_y, x=0, xanchor="left"),
        height=height,
        margin=dict(l=10, r=margin_right, t=margin_top, b=30),
        legend=dict(
            orientation="h",
            yanchor=legend_yanchor,
            y=legend_y,
            xanchor=legend_xanchor,
            x=legend_x,
            font=dict(size=legend_font_size),
        ),
    )
    if x_automargin:
        fig.update_xaxes(automargin=True)
    if dates and x_pad_days and x_pad_days > 0:
        min_x = min(dates) - pd.Timedelta(days=x_pad_days)
        max_x = max(dates) + pd.Timedelta(days=x_pad_days)
        fig.update_xaxes(range=[min_x, max_x])
    if counts:
        y_max = max(counts)
        if bar_headroom and bar_headroom > 0:
            y_max = y_max * (1 + bar_headroom)
        fig.update_yaxes(title_text="Completed tokens", secondary_y=False, range=[0, y_max])
    else:
        fig.update_yaxes(title_text="Completed tokens", secondary_y=False)
    tickvals, ticktext = build_time_ticks(max_tat)
    fig.update_yaxes(
        title_text="TAT (HH:MM)", secondary_y=True, tickvals=tickvals, ticktext=ticktext
    )
    return fig


def build_tat_distribution_chart(modality_code, modality_label):
    stage_avgs, stage_percentages, completed_count = compute_tat_distribution(modality_code)
    labels = list(stage_percentages.keys())
    values = [stage_percentages[l] for l in labels]
    label_text = [f"{label} {minutes_to_hhmm(stage_avgs.get(label, 0))}" for label in labels]
    color_map = {
        "Dispatch": "#1f77b4",
        "Billing": "#d62728",
        "Service": "#2ca02c",
    }
    colors = [color_map.get(label, "#999999") for label in labels]
    fig = go.Figure(
        data=[
            go.Pie(
                labels=label_text,
                values=values,
                hole=0.55,
                textinfo="percent",
                textposition="inside",
                marker=dict(colors=colors),
                hovertemplate="%{label}: %{percent}<extra></extra>",
            )
        ]
    )
    fig.update_layout(
        title=f"TAT Distribution by Process - {modality_label}",
        height=220,
        margin=dict(l=8, r=8, t=32, b=8),
        showlegend=True,
    )
    return fig, stage_avgs, stage_percentages, completed_count


def compute_us_service_breakdown():
    cols = [
        "Date",
        "Modality",
        "Status",
        ".ULTRASOUND - Wait Time",
        ".ULTRASOUND - Service Time",
        ".ULTRASOUND - Hold Time",
    ]
    df = read_filtered_xlsx(completed_path, cols)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df[(df["Date"] >= dt.datetime(2026, 1, 1)) & (df["Date"] < dt.datetime(2026, 2, 1))]
    df["Modality"] = df["Modality"].astype(str).str.strip().str.upper()
    df = df[df["Modality"] == "US"]
    if "Status" in df.columns:
        df["Status"] = normalize_status(df["Status"])
        df = df[df["Status"].isin(["Complete", "E-Complete"])]

    wait_m = df[".ULTRASOUND - Wait Time"].apply(to_minutes) if ".ULTRASOUND - Wait Time" in df.columns else pd.Series(dtype=float)
    service_m = df[".ULTRASOUND - Service Time"].apply(to_minutes) if ".ULTRASOUND - Service Time" in df.columns else pd.Series(dtype=float)
    hold_m = df[".ULTRASOUND - Hold Time"].apply(to_minutes) if ".ULTRASOUND - Hold Time" in df.columns else pd.Series(dtype=float)

    avg_wait = wait_m.mean() if len(wait_m) else 0
    avg_service = service_m.mean() if len(service_m) else 0
    avg_hold = hold_m.mean() if len(hold_m) else 0

    total = np.nansum([avg_wait, avg_service, avg_hold])
    if total <= 0:
        pct_wait = pct_service = pct_hold = 0
    else:
        pct_wait = (avg_wait / total) * 100
        pct_service = (avg_service / total) * 100
        pct_hold = (avg_hold / total) * 100

    avgs = {
        "Wait Time": 0 if pd.isna(avg_wait) else avg_wait,
        "Service Time": 0 if pd.isna(avg_service) else avg_service,
        "Hold Time": 0 if pd.isna(avg_hold) else avg_hold,
    }
    pcts = {
        "Wait Time": 0 if pd.isna(pct_wait) else pct_wait,
        "Service Time": 0 if pd.isna(pct_service) else pct_service,
        "Hold Time": 0 if pd.isna(pct_hold) else pct_hold,
    }
    return avgs, pcts


def compute_mri_service_breakdown():
    cols = [
        "Date",
        "Modality",
        "Status",
        ".MRI - Wait Time",
        ".MRI - Service Time",
        ".MRI - Hold Time",
    ]
    df = read_filtered_xlsx(completed_path, cols)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df[(df["Date"] >= dt.datetime(2026, 1, 1)) & (df["Date"] < dt.datetime(2026, 2, 1))]
    df["Modality"] = df["Modality"].astype(str).str.strip().str.upper()
    df = df[df["Modality"] == "MR"]
    if "Status" in df.columns:
        df["Status"] = normalize_status(df["Status"])
        df = df[df["Status"].isin(["Complete", "E-Complete"])]

    wait_m = df[".MRI - Wait Time"].apply(to_minutes) if ".MRI - Wait Time" in df.columns else pd.Series(dtype=float)
    service_m = df[".MRI - Service Time"].apply(to_minutes) if ".MRI - Service Time" in df.columns else pd.Series(dtype=float)
    hold_m = df[".MRI - Hold Time"].apply(to_minutes) if ".MRI - Hold Time" in df.columns else pd.Series(dtype=float)

    avg_wait = wait_m.mean() if len(wait_m) else 0
    avg_service = service_m.mean() if len(service_m) else 0
    avg_hold = hold_m.mean() if len(hold_m) else 0

    total = np.nansum([avg_wait, avg_service, avg_hold])
    if total <= 0:
        pct_wait = pct_service = pct_hold = 0
    else:
        pct_wait = (avg_wait / total) * 100
        pct_service = (avg_service / total) * 100
        pct_hold = (avg_hold / total) * 100

    avgs = {
        "Wait Time": 0 if pd.isna(avg_wait) else avg_wait,
        "Service Time": 0 if pd.isna(avg_service) else avg_service,
        "Hold Time": 0 if pd.isna(avg_hold) else avg_hold,
    }
    pcts = {
        "Wait Time": 0 if pd.isna(pct_wait) else pct_wait,
        "Service Time": 0 if pd.isna(pct_service) else pct_service,
        "Hold Time": 0 if pd.isna(pct_hold) else pct_hold,
    }
    return avgs, pcts


def build_us_service_breakdown_chart():
    avgs, pcts = compute_us_service_breakdown()
    labels = ["Wait Time", "Service Time", "Hold Time"]
    values = [pcts[l] for l in labels]
    label_text = [f"{l} {minutes_to_hhmm(avgs.get(l, 0))}" for l in labels]
    fig = go.Figure(
        data=[
            go.Pie(
                labels=label_text,
                values=values,
                hole=0.55,
                textinfo="percent",
                textposition="inside",
                marker=dict(colors=["#6A5ACD", "#E91E63", "#4FC3F7"]),
                hovertemplate="%{label}: %{percent}<extra></extra>",
            )
        ]
    )
    fig.update_traces(marker=dict(colors=["#6A5ACD", "#E91E63", "#4FC3F7"]))
    fig.update_layout(
        title="Ultrasound Service Breakdown",
        height=200,
        margin=dict(l=8, r=8, t=32, b=8),
        showlegend=True,
    )
    return fig, avgs, pcts


def build_mri_service_breakdown_chart():
    avgs, pcts = compute_mri_service_breakdown()
    labels = ["Wait Time", "Service Time", "Hold Time"]
    values = [pcts[l] for l in labels]
    label_text = [f"{l} {minutes_to_hhmm(avgs.get(l, 0))}" for l in labels]
    fig = go.Figure(
        data=[
            go.Pie(
                labels=label_text,
                values=values,
                hole=0.55,
                textinfo="percent",
                textposition="inside",
                marker=dict(colors=["#6A5ACD", "#E91E63", "#4FC3F7"]),
                hovertemplate="%{label}: %{percent}<extra></extra>",
            )
        ]
    )
    fig.update_layout(
        title="MRI Service Breakdown",
        height=200,
        margin=dict(l=8, r=8, t=32, b=8),
        showlegend=True,
    )
    return fig, avgs, pcts


scorecard = pd.read_csv(scorecard_path)
mod_status = pd.read_csv(modality_status_path)
mod_tat = pd.read_csv(modality_tat_path)

# Scorecards
scorecard_vals = scorecard.copy()
scorecard_vals["value"] = scorecard_vals["value"].astype(float)
score_fig = make_subplots(rows=2, cols=4, specs=[[{"type": "indicator"}] * 4] * 2)

labels = scorecard_vals["metric"].tolist()
values = scorecard_vals["value"].tolist()

for i, (label, value) in enumerate(zip(labels, values)):
    r = i // 4 + 1
    c = i % 4 + 1
    display_value = value
    if label == "Completion rate":
        display_value = value * 100
        number = {"suffix": "%", "valueformat": ".1f"}
    else:
        number = {"valueformat": ",.0f"}
    score_fig.add_trace(
        go.Indicator(mode="number", value=display_value, title={"text": label}, number=number),
        row=r,
        col=c,
    )

score_fig.update_layout(height=350, margin=dict(l=20, r=20, t=20, b=10))

# Modality status chart
mod_status_fig = go.Figure()
mod_status_fig.add_trace(
    go.Bar(
        x=mod_status["Modality"],
        y=mod_status["Completed tokens"],
        name="Completed",
        marker_color="#2ca02c",
    )
)
mod_status_fig.add_trace(
    go.Bar(
        x=mod_status["Modality"],
        y=mod_status["Non-completed tokens"],
        name="Non-completed",
        marker_color="#d62728",
    )
)
mod_status_fig.update_layout(
    barmode="group",
    height=360,
    margin=dict(l=20, r=20, t=30, b=20),
    title="Completed vs Non-completed Tokens by Modality",
)

# Modality TAT vs Target
focus = ["US", "MR", "XR", "CT"]
mod_tat_focus = mod_tat[mod_tat["Modality"].isin(focus)].copy()
mod_map = {"US": "Ultrasound", "MR": "MRI", "XR": "XRAY", "CT": "CT"}
mod_tat_focus["Modality"] = mod_tat_focus["Modality"].map(mod_map)

mod_tat_fig = go.Figure()
mod_tat_fig.add_trace(
    go.Bar(
        x=mod_tat_focus["Modality"],
        y=mod_tat_focus["mean_actual_min"],
        name="Average TAT",
        marker_color="#636efa",
        text=[minutes_to_hhmm(v) for v in mod_tat_focus["mean_actual_min"]],
        textposition="outside",
        hovertemplate="%{x}: %{customdata}",
    )
)
mod_tat_fig.data[-1].customdata = [minutes_to_hhmm(v) for v in mod_tat_focus["mean_actual_min"]]
mod_tat_fig.add_trace(
    go.Bar(
        x=mod_tat_focus["Modality"],
        y=mod_tat_focus["target_min"],
        name="Target TAT",
        marker_color="#00cc96",
        text=[minutes_to_hhmm(v) for v in mod_tat_focus["target_min"]],
        textposition="outside",
        hovertemplate="%{x}: %{customdata}",
    )
)
mod_tat_fig.data[-1].customdata = [minutes_to_hhmm(v) for v in mod_tat_focus["target_min"]]
mod_tat_fig.update_layout(
    barmode="group",
    height=360,
    margin=dict(l=20, r=20, t=30, b=20),
    title="Average TAT vs Target by Modality",
)
mod_max = np.nanmax([mod_tat_focus["mean_actual_min"].max(), mod_tat_focus["target_min"].max()])
mod_tickvals, mod_ticktext = build_time_ticks(mod_max)
mod_tat_fig.update_yaxes(title_text="TAT (HH:MM)", tickvals=mod_tickvals, ticktext=mod_ticktext)

# Daily combo charts by modality
combo_xray = build_combo_chart(
    "XR",
    "XRAY",
    height=380,
    legend_y=1.02,
    legend_x=0.5,
    legend_xanchor="center",
    legend_yanchor="bottom",
    margin_top=120,
    legend_font_size=10,
    bar_headroom=0.60,
    bar_cliponaxis=False,
)
combo_mri = build_combo_chart(
    "MR",
    "MRI",
    height=300,
    legend_y=1.02,
    legend_x=0.5,
    legend_xanchor="center",
    legend_yanchor="bottom",
    margin_top=95,
    legend_font_size=10,
    bar_headroom=0.12,
)
combo_ct = build_combo_chart(
    "CT",
    "CT",
    height=380,
    legend_y=1.02,
    legend_x=0.5,
    legend_xanchor="center",
    legend_yanchor="bottom",
    margin_top=120,
    legend_font_size=10,
    bar_headroom=0.60,
    bar_cliponaxis=False,
)
combo_us = build_combo_chart(
    "US",
    "Ultrasound",
    height=580,
    legend_y=1.02,
    title_y=1.0,
    legend_x=0.5,
    legend_xanchor="center",
    legend_yanchor="bottom",
    margin_top=90,
    legend_font_size=10,
    bar_headroom=0.06,
)

tat_dist_xray, tat_avgs_xray, tat_pct_xray, tat_count_xray = build_tat_distribution_chart("XR", "XRAY")
tat_dist_mri, tat_avgs_mri, tat_pct_mri, tat_count_mri = build_tat_distribution_chart("MR", "MRI")
tat_dist_ct, tat_avgs_ct, tat_pct_ct, tat_count_ct = build_tat_distribution_chart("CT", "CT SCAN")
tat_dist_us, tat_avgs_us, tat_pct_us, tat_count_us = build_tat_distribution_chart("US", "ULTRASOUND")
us_service_breakdown, us_service_avgs, us_service_pcts = build_us_service_breakdown_chart()
mri_service_breakdown, mri_service_avgs, mri_service_pcts = build_mri_service_breakdown_chart()

def build_distribution_table(stage_avgs, stage_percentages):
    rows = []
    for stage in ["Billing", "Service", "Dispatch"]:
        avg_val = minutes_to_hhmm(stage_avgs.get(stage, 0))
        pct_val = stage_percentages.get(stage, 0)
        rows.append(f"<tr><td>{stage}</td><td>{avg_val}</td><td>{pct_val:.1f}%</td></tr>")
    return (
        "<table class=\"mini-table\">"
        "<thead><tr><th>Process</th><th>Avg TAT (HH:MM)</th><th>% of Total</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table>"
    )


def build_service_breakdown_table(stage_avgs, stage_percentages):
    rows = []
    for stage in ["Wait Time", "Service Time", "Hold Time"]:
        avg_val = minutes_to_hhmm(stage_avgs.get(stage, 0))
        pct_val = stage_percentages.get(stage, 0)
        rows.append(f"<tr><td>{stage}</td><td>{avg_val}</td><td>{pct_val:.1f}%</td></tr>")
    return (
        "<table class=\"mini-table\">"
        "<thead><tr><th>Component</th><th>Avg (HH:MM)</th><th>% of Service</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table>"
    )

dist_table_xray = build_distribution_table(tat_avgs_xray, tat_pct_xray)
dist_table_mri = build_distribution_table(tat_avgs_mri, tat_pct_mri)
dist_table_ct = build_distribution_table(tat_avgs_ct, tat_pct_ct)
dist_table_us = build_distribution_table(tat_avgs_us, tat_pct_us)
us_service_table = build_service_breakdown_table(us_service_avgs, us_service_pcts)
mri_service_table = build_service_breakdown_table(mri_service_avgs, mri_service_pcts)

html = f"""
<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"UTF-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\" />
  <title>January 2026 TAT & Token Workflow Report</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 0; background: #f7f7f9; color: #222; }}
    header {{ background: #0b3d91; color: #fff; padding: 16px 32px; display: flex; align-items: center; justify-content: space-between; gap: 16px; }}
    header img {{ height: 44px; width: auto; background: #fff; border-radius: 6px; padding: 4px; }}
    h1 {{ margin: 0; font-size: 22px; }}
    .container {{ padding: 20px 28px 40px; }}
    .card {{ background: #fff; border-radius: 8px; padding: 16px; margin-bottom: 18px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
    .grid {{ display: grid; grid-template-columns: 1.6fr 1fr; gap: 12px; align-items: stretch; }}
    .left-stack {{ display: flex; flex-direction: column; height: 100%; }}
    .right-stack {{ display: flex; flex-direction: column; gap: 8px; height: 100%; }}
    .bullets {{ margin: 8px 0 0 0; padding-left: 18px; }}
    .bullets li {{ margin: 4px 0; }}
    .mini-table {{ width: 100%; border-collapse: collapse; margin-top: 4px; font-size: 11.5px; }}
    .mini-table th, .mini-table td {{ border: 1px solid #ddd; padding: 4px 6px; text-align: left; }}
    .mini-table th {{ background: #f0f3f8; }}
    .note {{ font-size: 11.5px; color: #555; margin-top: 4px; }}
    @media (max-width: 1100px) {{ .grid {{ grid-template-columns: 1fr; }} }}
  </style>
  <script src=\"https://cdn.plot.ly/plotly-2.32.0.min.js\"></script>
</head>
<body>
  <header>
    <h1>January 2026 TAT & Token Workflow Report</h1>
    <img src=\"{logo_src}\" alt=\"Sonar Imaging Centre\" />
  </header>
  <div class=\"container\">
    <div class=\"card\">
      <div id=\"scorecards\"></div>
      <ul class=\"bullets\">
        <li>Monthly intake volume is high with a completion rate of 68.6%.</li>
      </ul>
    </div>

    <div class=\"card\">
      <div id=\"modality_status\"></div>
      <ul class=\"bullets\">
        <li>MRI has the largest incomplete workload; XRAY shows moderate incomplete load.</li>
        <li>Prioritize MRI throughput to stabilize overall completion.</li>
      </ul>
    </div>

    <div class=\"card\">
      <div id=\"tat_vs_target\"></div>
      <ul class=\"bullets\">
        <li>All four modalities exceed target TAT on average.</li>
                <li>MRI shows the widest gap.</li>
      </ul>
    </div>

        <div class=\"card\">
            <div class=\"grid\">
                <div class=\"left-stack\">
                    <div id=\"combo_xray\"></div>
                    <ul class=\"bullets\">
                        <li>Daily XRAY completed volume vs average TAT with target reference.</li>
                    </ul>
                    <ul class=\"bullets\">
                        <li>XRAY TAT fluctuated across January and was above target on several days. Reported TAT reflects completed tokens only; spine patients are excluded because their workflow time is not fully visible in the system.</li>
                    </ul>
                </div>
                <div class=\"right-stack\">
                    <div id=\"tat_dist_xray\"></div>
                    {dist_table_xray}
                </div>
            </div>
        </div>

        <div class=\"card\">
            <div class=\"grid\">
                <div class=\"left-stack\">
                    <div id=\"combo_mri\"></div>
                    <ul class=\"bullets\">
                        <li>Daily MRI completed volume vs average TAT with target reference.</li>
                    </ul>
                    <ul class=\"bullets\">
                        <li>MRI TAT remained above target on most days throughout the month. Reported TAT reflects completed tokens only; For tokens where service was done but were left pending or standby status are excluded, as they did not complete the full cycle.</li>
                        <li>MRI Service TAT (completed tokens only) is predominantly driven by wait time. The longest recorded wait was 06:58, the shortest 00:01, with a median wait of 01:14. Wait time contributes approximately 77.5% of total Service TAT, indicating that delays are primarily driven by pre-service queueing rather than scan duration.</li>
                    </ul>
                </div>
                <div class=\"right-stack\">
                    <div id=\"tat_dist_mri\"></div>
                    {dist_table_mri}
                    <div id=\"mri_service_breakdown\"></div>
                    {mri_service_table}
                    <div class=\"note\">Wait Time = time before patient is called; Service Time = time in service; Hold Time = preparation time.</div>
                </div>
            </div>
        </div>

        <div class=\"card\">
            <div class=\"grid\">
                <div class=\"left-stack\">
                    <div id=\"combo_ct\"></div>
                    <ul class=\"bullets\">
                        <li>Daily CT completed volume vs average TAT with target reference.</li>
                    </ul>
                    <ul class=\"bullets\">
                        <li>CT TAT was above the target on most days in January, with a few days showing high spikes above 4 hours, indicating that longer turnaround times occurred on specific days rather than consistently throughout the month.</li>
                        <li>Most of the CT TAT is driven by dispatch (about 54%), while the actual scan time is relatively short.</li>
                    </ul>
                </div>
                <div class=\"right-stack\">
                    <div id=\"tat_dist_ct\"></div>
                    {dist_table_ct}
                </div>
            </div>
        </div>

        <div class=\"card\">
            <div class=\"grid\">
                <div class=\"left-stack\">
                    <div id=\"combo_us\"></div>
                    <ul class=\"bullets\">
                        <li>Daily Ultrasound completed volume vs average TAT with target reference.</li>
                    </ul>
                    <ul class=\"bullets\">
                        <li>Ultrasound TAT shows significant day-to-day variability and frequently exceeds target, with the highest TAT exceeding 3 hours.</li>
                        <li>Wait time is the dominant contributor to service TAT (longest 05:06, median 00:48), driven primarily by arrival surges and patient readiness before service rather than scan duration.</li>
                        <li>Strengthening patient preparation and readiness workflows prior to service would reduce queue buildup (wait time) and improve overall Ultrasound TAT performance.</li>
                    </ul>
                </div>
                <div class=\"right-stack\">
                    <div id=\"tat_dist_us\"></div>
                    {dist_table_us}
                    <div id=\"us_service_breakdown\"></div>
                    {us_service_table}
                    <div class=\"note\">Wait Time = time before patient is called; Service Time = time in service; Hold Time = preparation time.</div>
                </div>
            </div>
        </div>
  </div>

  <script>
    var scoreData = {score_fig.to_json()};
    var modStatusData = {mod_status_fig.to_json()};
    var tatVsTargetData = {mod_tat_fig.to_json()};
    var comboXrayData = {combo_xray.to_json()};
    var comboMriData = {combo_mri.to_json()};
    var comboCtData = {combo_ct.to_json()};
    var comboUsData = {combo_us.to_json()};
    var tatDistXrayData = {tat_dist_xray.to_json()};
    var tatDistMriData = {tat_dist_mri.to_json()};
    var tatDistCtData = {tat_dist_ct.to_json()};
    var tatDistUsData = {tat_dist_us.to_json()};
    var usServiceBreakdownData = {us_service_breakdown.to_json()};
    var mriServiceBreakdownData = {mri_service_breakdown.to_json()};

    Plotly.newPlot('scorecards', scoreData.data, scoreData.layout, {{displayModeBar: false}});
    Plotly.newPlot('modality_status', modStatusData.data, modStatusData.layout, {{displayModeBar: false}});
    Plotly.newPlot('tat_vs_target', tatVsTargetData.data, tatVsTargetData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_xray', comboXrayData.data, comboXrayData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_mri', comboMriData.data, comboMriData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_ct', comboCtData.data, comboCtData.layout, {{displayModeBar: false}});
    Plotly.newPlot('combo_us', comboUsData.data, comboUsData.layout, {{displayModeBar: false}});
    Plotly.newPlot('tat_dist_xray', tatDistXrayData.data, tatDistXrayData.layout, {{displayModeBar: false}});
    Plotly.newPlot('tat_dist_mri', tatDistMriData.data, tatDistMriData.layout, {{displayModeBar: false}});
    Plotly.newPlot('tat_dist_ct', tatDistCtData.data, tatDistCtData.layout, {{displayModeBar: false}});
    Plotly.newPlot('tat_dist_us', tatDistUsData.data, tatDistUsData.layout, {{displayModeBar: false}});
    Plotly.newPlot('us_service_breakdown', usServiceBreakdownData.data, usServiceBreakdownData.layout, {{displayModeBar: false}});
    Plotly.newPlot('mri_service_breakdown', mriServiceBreakdownData.data, mriServiceBreakdownData.layout, {{displayModeBar: false}});
  </script>
</body>
</html>
"""

with open(os.path.join(base, "dashboard.html"), "w", encoding="utf-8") as f:
    f.write(html)
