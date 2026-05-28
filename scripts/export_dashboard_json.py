from __future__ import annotations

import argparse
import json
import math
import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR   = Path(__file__).resolve().parents[1]
INPUT_DIR  = BASE_DIR / "data" / "input"
OUTPUT_DIR = BASE_DIR / "docs" / "data"          # ← JSON lands here

ALL_TOKENS_PATH       = INPUT_DIR / "TAT - ALL TOKENS.xlsx"
COMPLETED_TOKENS_PATH = INPUT_DIR / "TAT - ALL COMPLETED TOKENS.xlsx"

DATA_SHEET = "DATA Daily Token Service Report"

# ---------------------------------------------------------------------------
# Lookup tables
# ---------------------------------------------------------------------------
MODALITY_LABELS = {
    "XR": "XRAY",
    "MR": "MRI",
    "CT": "CT",
    "US": "Ultrasound",
}

# Statuses that count as "Completed" for scorecards and completion rate.
# FIX: E-Complete is a genuine completed status and is now included everywhere.
COMPLETED_STATUSES = {"Complete"}

# Regex matching any status that ends with " - Done" (case-insensitive).
# Used to classify Non-Complete Done tokens.
_DONE_SUFFIX_RE = re.compile(r" - Done$", re.IGNORECASE)

# Statuses that end in "- Done" but are EXCLUDED from the Non-Complete Done bucket
# (they fall into the plain Non-Complete bucket instead).
# ─── TO ADD NEW EXCLUSIONS: just extend this set ───────────────────────────
NON_COMPLETE_DONE_EXCLUDE: set[str] = {"Complete - Done"}


# ── Status bucket helpers ────────────────────────────────────────────────────
# These three predicates drive all three-bucket logic across the codebase.
# Adding a new status requires no code changes — just update the sets above.

def is_completed(status: str) -> bool:
    """Complete or E-Complete."""
    return status in COMPLETED_STATUSES


def is_non_complete_done(status: str) -> bool:
    """Ends with ' - Done' and is NOT in the exclusion set."""
    return bool(_DONE_SUFFIX_RE.search(status)) and status not in NON_COMPLETE_DONE_EXCLUDE


def is_non_complete(status: str) -> bool:
    """Everything that is neither completed nor non-complete-done."""
    return not is_completed(status) and not is_non_complete_done(status)


ALL_TOKEN_ALIASES = {
    "date":     ["Date"],
    "token":    ["Token"],
    "modality": ["Modality"],
    "status":   ["Status"],
}

COMPLETED_TOKEN_ALIASES = {
    "date":           ["Date"],
    "token":          ["Token"],
    "modality":       ["Modality"],
    "status":         ["Status"],
    "actual_tat":     ["ACTUAL Turnaround Time"],
    "target_tat":     ["TARGET TAT"],
    "dispatch_tat":   ["TAT - DISPATCH SERVICE"],
    "us_service_tat": ["TAT - US SERVICE", "U/S TAT"],
    "xr_service_tat": ["XRAY - TAT SERVICE", "Wait Time + Service time"],
    "ct_service_tat": ["TAT - CT SERVICE", "CT TAT"],
    "mr_service_tat": ["TAT - MRI SEVICE", "TAT - MRI SERVICE", "MRI TAT"],
    "us_billing_tat": ["TAT - U/S BILLING", "US BILLING TAT"],
    "xr_billing_tat": ["TAT - BILLING XRAY", "XRAY BILLING TAT"],
    "ct_billing_tat": ["TAT - CT BILLING", "CT BILLING TAT"],
    "mr_billing_tat": ["TAT - MRI BILLING", "MRI BILLING TAT"],
    "us_stage_wait":     [".ULTRASOUND - Wait Time"],
    "us_stage_service":  [".ULTRASOUND - Service Time"],
    "us_stage_hold":     [".ULTRASOUND - Hold Time"],
    "xr_stage_wait":     ["X-RAY - Wait Time"],
    "xr_stage_service":  ["X-RAY - Service Time"],
    "xr_stage_hold":     ["X-RAY - Hold Time"],
    "ct_stage_wait":     [".CT SCAN - Wait Time"],
    "ct_stage_service":  [".CT SCAN - Service Time"],
    "ct_stage_hold":     [".CT SCAN - Hold Time"],
    "mr_stage_wait":     [".MRI - Wait Time"],
    "mr_stage_service":  [".MRI - Service Time"],
    "mr_stage_hold":     [".MRI - Hold Time"],
    "us_billing_wait":    ["ULTRASOUND SERVICE - Wait Time"],
    "us_billing_service": ["ULTRASOUND SERVICE - Service Time"],
    "us_billing_hold":    ["ULTRASOUND SERVICE - Hold Time"],
    "xr_billing_wait":    ["X-RAY SERVICE - Wait Time"],
    "xr_billing_service": ["X-RAY SERVICE - Service Time"],
    "xr_billing_hold":    ["X-RAY SERVICE - Hold Time"],
    "ct_billing_wait":    ["CT SCAN SERVICE - Wait Time"],
    "ct_billing_service": ["CT SCAN SERVICE - Service Time"],
    "ct_billing_hold":    ["CT SCAN SERVICE - Hold Time"],
    "mr_billing_wait":    ["MRI SERVICE - Wait Time"],
    "mr_billing_service": ["MRI SERVICE - Service Time"],
    "mr_billing_hold":    ["MRI SERVICE - Hold Time"],
}

SERVICE_COLUMNS_BY_MODALITY = {
    "US": "us_service_tat",
    "XR": "xr_service_tat",
    "CT": "ct_service_tat",
    "MR": "mr_service_tat",
}

BILLING_COLUMNS_BY_MODALITY = {
    "US": "us_billing_tat",
    "XR": "xr_billing_tat",
    "CT": "ct_billing_tat",
    "MR": "mr_billing_tat",
}

MODALITY_CODE_TO_LABEL = {v: k for k, v in MODALITY_LABELS.items()}


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def clean_column_name(value: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z]+", "_", str(value).strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned.lower()


def normalize_status(value: object) -> str:
    """
    Normalise a raw status string into a stable canonical form.

    Canonical forms produced (extensible — see sets at top of file):
        Complete, E-Complete
        Pending, Serving, Standby, Noshow
        <Base> - Done    (e.g. Standby - Done, Pending - Done)
        <Base> - Not Done
        Complete - Done  (excluded from Non-Complete Done bucket)
    """
    if value is None or pd.isna(value):
        return "Unknown"
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return "Unknown"

    # Normalise spacing around hyphens: any amount of spaces + hyphen + spaces → " - "
    text = re.sub(r"\s*-\s*", " - ", text)
    # Collapse any internal whitespace runs
    text = re.sub(r"\s+", " ", text).strip()

    # E. Complete / E Complete → E-Complete  (before title-casing so "e" stays lowercase)
    text = re.sub(r"E\.\s*Complete", "E-Complete", text, flags=re.IGNORECASE)
    text = re.sub(r"E Complete",     "E-Complete", text, flags=re.IGNORECASE)

    # No Show / No-show variants → Noshow
    text = re.sub(r"No[\s-]?[Ss]how", "Noshow", text)

    # Normalise DONE / NOT DONE suffix casing
    text = re.sub(r"\bDONE\b",     "Done",     text)
    text = re.sub(r"\bNOT DONE\b", "Not Done", text, flags=re.IGNORECASE)

    # Title-case only the base word (the part before the first " - ")
    parts = text.split(" - ", 1)
    parts[0] = parts[0].capitalize()
    text = " - ".join(parts)

    # Final hardcoded overrides for residual edge-cases
    overrides = {
        "Noshow":     "Noshow",
        "E-complete": "E-Complete",
        "E-Complete": "E-Complete",
    }
    return overrides.get(text, text)


def normalize_modality(value: object) -> str:
    if value is None or pd.isna(value):
        return "Unknown"
    code = str(value).strip().upper()
    if not code or code == "NAN":
        return "Unknown"
    return MODALITY_LABELS.get(code, "Unknown")


_UNKNOWN_MODALITY_LABELS: set[str] = set()   # collected; reported once at the end


def modality_code_from_label(label: str) -> str | None:
    """
    Return the modality code for a display label (e.g. 'XRAY' -> 'XR').
    Unknown labels are silently collected and reported as a single tidy
    summary at the end of the run, not as repetitive per-row warnings.
    """
    code = MODALITY_CODE_TO_LABEL.get(label)
    if code is None:
        _UNKNOWN_MODALITY_LABELS.add(label)
    return code


def to_minutes(value: object) -> float:
    """
    FIX: the previous heuristic treated any float between 0 and 2 as an Excel
    day-fraction, which incorrectly multiplied genuine 1-minute TAT values by
    1440. The corrected rule only applies the day-fraction conversion when the
    value is a non-integer float strictly between 0 and 1 (i.e. a true
    Excel serial fraction less than one full day).
    """
    if value is None:
        return math.nan
    if isinstance(value, float) and math.isnan(value):
        return math.nan
    if isinstance(value, (int, float)):
        number = float(value)
        # Only treat as Excel day-fraction if it's a non-integer between 0 and 1
        if isinstance(value, float) and 0 < number < 1:
            return number * 24 * 60
        return number
    if isinstance(value, pd.Timedelta):
        return value.total_seconds() / 60
    if isinstance(value, timedelta):
        return value.total_seconds() / 60
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, time):
        return value.hour * 60 + value.minute + value.second / 60
    text = str(value).strip()
    if text in {"", "--", "nan", "NaN", "None", "0:00:00"}:
        return math.nan
    try:
        delta = pd.to_timedelta(text)
        return delta.total_seconds() / 60
    except Exception:
        pass
    try:
        number = float(text)
        if isinstance(number, float) and 0 < number < 1:
            return number * 24 * 60
        return number
    except ValueError:
        return math.nan


def minutes_to_hhmm(value: float) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "00:00"
    total_minutes = max(int(round(float(value))), 0)
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours:02d}:{minutes:02d}"


def safe_number(value: float, digits: int = 2) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return round(float(value), digits)


def month_key(d: date) -> str:
    return d.strftime("%b").lower()


def month_label(d: date) -> str:
    return d.strftime("%B %Y")


# ---------------------------------------------------------------------------
# Excel loading
# ---------------------------------------------------------------------------

def load_selected_columns(
    workbook_path: Path,
    sheet_name: str,
    column_aliases: dict[str, list[str]],
) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header_cells = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None
        )
        if not header_cells:
            return pd.DataFrame(columns=sorted(column_aliases))

        normalized_headers = {
            normalize_text(header): index
            for index, header in enumerate(header_cells)
            if header is not None and str(header).strip()
        }

        selected_columns: dict[str, int] = {}
        for output_name, aliases in column_aliases.items():
            for alias in aliases:
                normalized_alias = normalize_text(alias)
                if normalized_alias in normalized_headers:
                    selected_columns[output_name] = normalized_headers[normalized_alias]
                    break

        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            record = {}
            for output_name, column_index in selected_columns.items():
                record[output_name] = row[column_index] if column_index < len(row) else None
            rows.append(record)

        frame = pd.DataFrame(rows)
        for output_name in column_aliases:
            if output_name not in frame.columns:
                frame[output_name] = pd.NA

        frame = frame[[column for column in column_aliases]]
        frame.columns = [clean_column_name(column) for column in frame.columns]
        return frame
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Data preparation
# ---------------------------------------------------------------------------

def prepare_all_tokens(path: Path) -> pd.DataFrame:
    frame = load_selected_columns(path, DATA_SHEET, ALL_TOKEN_ALIASES)
    frame["date"]     = pd.to_datetime(frame["date"], errors="coerce").dt.date
    frame["token"]    = frame["token"].fillna("").astype(str).str.strip()
    frame["modality"] = frame["modality"].apply(normalize_modality)
    frame["status"]   = frame["status"].apply(normalize_status)
    return frame


def prepare_completed_tokens(path: Path) -> pd.DataFrame:
    frame = load_selected_columns(path, DATA_SHEET, COMPLETED_TOKEN_ALIASES)
    frame["date"]     = pd.to_datetime(frame["date"], errors="coerce").dt.date
    frame["token"]    = frame["token"].fillna("").astype(str).str.strip()
    frame["modality"] = frame["modality"].apply(normalize_modality)
    frame["status"]   = frame["status"].apply(normalize_status)

    frame["actual_tat_minutes"]   = frame["actual_tat"].apply(to_minutes)
    frame["target_tat_minutes"]   = frame["target_tat"].apply(to_minutes)
    frame["dispatch_tat_minutes"] = frame["dispatch_tat"].apply(to_minutes)

    for modality, column in SERVICE_COLUMNS_BY_MODALITY.items():
        frame[f"{column}_minutes"] = frame[column].apply(to_minutes)
    for modality, column in BILLING_COLUMNS_BY_MODALITY.items():
        frame[f"{column}_minutes"] = frame[column].apply(to_minutes)

    component_columns = [
        "us_stage_wait", "us_stage_service", "us_stage_hold",
        "xr_stage_wait", "xr_stage_service", "xr_stage_hold",
        "ct_stage_wait", "ct_stage_service", "ct_stage_hold",
        "mr_stage_wait", "mr_stage_service", "mr_stage_hold",
        "us_billing_wait", "us_billing_service", "us_billing_hold",
        "xr_billing_wait", "xr_billing_service", "xr_billing_hold",
        "ct_billing_wait", "ct_billing_service", "ct_billing_hold",
        "mr_billing_wait", "mr_billing_service", "mr_billing_hold",
    ]
    for column in component_columns:
        frame[f"{column}_minutes"] = frame[column].apply(to_minutes)

    return frame


# ---------------------------------------------------------------------------
# Builders
# ---------------------------------------------------------------------------

def _months_present(frame: pd.DataFrame) -> list[date]:
    dates = frame["date"].dropna().unique()
    seen: dict[tuple[int, int], date] = {}
    for d in sorted(dates):
        key = (d.year, d.month)
        if key not in seen:
            seen[key] = d
    return list(seen.values())


def build_scorecards(all_tokens: pd.DataFrame) -> dict:
    """
    Per-month scorecard summary.

    Each status bucket (pending, serving, standby, noshow, e-complete)
    includes all variants of that status — e.g. pendingTokens counts
    Pending, Pending - Done, Pending - NOT Done, Pending -Done, etc.

    NOTE: completedTokens is intentionally an exact match for 'Complete'
    only — 'Complete - Done' belongs in the nonCompleteDone bucket.
    """
    def sum_by_base(status_counts, base: str) -> int:
        """
        Sum all statuses whose base word matches, regardless of suffix variants.
        Safely handles E-Complete whose internal hyphen would otherwise be
        mistaken for a bucket separator.
        """
        base_lower = base.strip().lower()
        total = 0
        for status, count in status_counts.items():
            s = str(status).strip()
            # Step 1: protect E-Complete's internal hyphen BEFORE normalizing separators
            # Covers: E-Complete, E. Complete, E .Complete etc (post-normalize_status form)
            s = re.sub(r'\bE\s*[-\.]\s*Complete\b', 'ECOMPLETE_PLACEHOLDER', s, flags=re.IGNORECASE)
            # Step 2: normalize all remaining ' - ' separators
            s = re.sub(r'\s*-\s*', ' - ', s)
            # Step 3: restore placeholder
            s = s.replace('ECOMPLETE_PLACEHOLDER', 'E-Complete')
            # Step 4: extract base (everything before first ' - ')
            base_part = s.split(' - ')[0].strip().lower()
            if base_part == base_lower:
                total += int(count)
        return total

    months = _months_present(all_tokens)
    result = []
    for rep in months:
        key   = month_key(rep)
        label = month_label(rep)
        subset = all_tokens[
            all_tokens["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        status_counts   = subset["status"].value_counts()

        # completedTokens = exact 'Complete' only (Complete - Done goes to nonCompleteDone)
        completed_total = int(status_counts.get("Complete",   0))
        # eCompleteTokens = all E-Complete variants (E-Complete, E-Complete - Done, etc.)
        ecomplete_total = sum_by_base(status_counts, "E-Complete")
        all_completed   = completed_total 
        total           = int(len(subset))

        result.append({
            "key":                key,
            "label":              label,
            "totalTokens":        total,
            "completedTokens":    completed_total,
            "eCompleteTokens":    ecomplete_total,
            "allCompletedTokens": all_completed,
            "pendingTokens":      sum_by_base(status_counts, "Pending"),
            "servingTokens":      sum_by_base(status_counts, "Serving"),
            "noShowTokens":       sum_by_base(status_counts, "Noshow"),
            "standbyTokens":      sum_by_base(status_counts, "Standby"),
            "completionRate":     safe_number(
                (all_completed / total * 100) if total else 0.0, 1
            ),
        })
    return {"months": result}

def build_modality_status(all_tokens: pd.DataFrame) -> dict:
    """
    Per-month, per-modality status breakdown into THREE buckets:

      completed        — Complete + E-Complete
      nonCompleteDone  — any status ending in ' - Done' except 'Complete - Done'
      nonCompleted     — everything else

    The raw statusBreakdown dict is also emitted so the frontend can show
    per-status counts without touching Python.  Adding a new status to the
    data requires no code changes — just update COMPLETED_STATUSES or
    NON_COMPLETE_DONE_EXCLUDE at the top of the file.

    FIX (was bug): E-Complete was incorrectly going into nonCompleted.
    """
    months  = _months_present(all_tokens)
    cleaned = all_tokens[all_tokens["modality"] != "Unknown"].copy()
    result  = []
    for rep in months:
        subset = cleaned[
            cleaned["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby(["modality", "status"], dropna=False)
            .size()
            .unstack(fill_value=0)
            .sort_index()
        )
        modalities = []
        for modality, row in grouped.iterrows():
            completed_count     = sum(int(row.get(s, 0)) for s in row.index if is_completed(s))
            nc_done_count       = sum(int(row.get(s, 0)) for s in row.index if is_non_complete_done(s))
            nc_count            = sum(int(row.get(s, 0)) for s in row.index if is_non_complete(s))
            status_breakdown    = {
                clean_column_name(s): int(c)
                for s, c in row.items() if int(c) > 0
            }
            modalities.append({
                "modality":        modality,
                "completed":       completed_count,
                "nonCompleteDone": nc_done_count,
                "nonCompleted":    nc_count,
                "statusBreakdown": status_breakdown,
            })
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


def build_tat_vs_target(completed_tokens: pd.DataFrame) -> dict:
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        (completed_tokens["modality"] != "Unknown")
    ].copy()
    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby("modality", dropna=False)
            .agg(
                actual_minutes=("actual_tat_minutes", "mean"),
                target_minutes=("target_tat_minutes", "mean"),
                token_count=("token", "count"),
            )
            .reset_index()
            .sort_values("modality")
        )
        modalities = []
        for row in grouped.itertuples(index=False):
            modalities.append({
                "modality":      row.modality,
                "actualMinutes": safe_number(row.actual_minutes),
                "targetMinutes": safe_number(row.target_minutes),
                "actualHHMM":   minutes_to_hhmm(row.actual_minutes),
                "targetHHMM":   minutes_to_hhmm(row.target_minutes),
                "tokenCount":   int(row.token_count),
            })
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


def build_tat_distribution(completed_tokens: pd.DataFrame) -> dict:
    """
    Per-month TAT split across billing / service / dispatch by modality.

    FIX 1 — Chart vs Table % mismatch: percentages are now computed from the
    same rounded minute values written to JSON, so Plotly's own percentage
    calculation (which also uses those values) will always match the table.

    FIX 2 — Partial totals: when any component is NaN it is excluded from the
    total so the remaining percentages still sum to 100 % and are coherent.
    """
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        (completed_tokens["modality"] != "Unknown")
    ].copy()
    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        modalities = []
        for modality in sorted(subset["modality"].dropna().unique()):
            modality_code  = modality_code_from_label(modality)
            if modality_code is None:
                continue
            service_column = SERVICE_COLUMNS_BY_MODALITY.get(modality_code)
            billing_column = BILLING_COLUMNS_BY_MODALITY.get(modality_code)
            if not service_column or not billing_column:
                continue
            mf = subset[subset["modality"] == modality]

            billing_rounded  = safe_number(mf[f"{billing_column}_minutes"].mean())
            service_rounded  = safe_number(mf[f"{service_column}_minutes"].mean())
            dispatch_rounded = safe_number(mf["dispatch_tat_minutes"].mean())

            # Compute % from the SAME rounded values Plotly will receive
            total_rounded = sum(
                v for v in [billing_rounded, service_rounded, dispatch_rounded]
                if v is not None
            )

            def pct(v: float | None) -> float | None:
                if v is None or total_rounded == 0:
                    return None
                return round(v / total_rounded * 100, 1)

            modalities.append({
                "modality":        modality,
                "billingMinutes":  billing_rounded,
                "billingHHMM":     minutes_to_hhmm(mf[f"{billing_column}_minutes"].mean()),
                "billingPct":      pct(billing_rounded),
                "serviceMinutes":  service_rounded,
                "serviceHHMM":     minutes_to_hhmm(mf[f"{service_column}_minutes"].mean()),
                "servicePct":      pct(service_rounded),
                "dispatchMinutes": dispatch_rounded,
                "dispatchHHMM":    minutes_to_hhmm(mf["dispatch_tat_minutes"].mean()),
                "dispatchPct":     pct(dispatch_rounded),
                "tokenCount":      int(len(mf)),
            })
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


def build_daily_trends(completed_tokens: pd.DataFrame) -> dict:
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        completed_tokens["date"].notna() &
        (completed_tokens["modality"] != "Unknown")
    ].copy()

    label_to_key = {v: k for k, v in MODALITY_LABELS.items()}

    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby(["modality", "date"], dropna=False)
            .agg(
                tokens=("token", "count"),
                actual_minutes=("actual_tat_minutes", "mean"),
                target_minutes=("target_tat_minutes", "mean"),
            )
            .reset_index()
            .sort_values(["modality", "date"])
        )
        modalities_dict: dict[str, list] = {}
        for modality, mf in grouped.groupby("modality", sort=True):
            js_key = label_to_key.get(modality, modality)
            points = []
            for row in mf.itertuples(index=False):
                points.append({
                    "date":       row.date.isoformat() if isinstance(row.date, date) else str(row.date),
                    "completed":  int(row.tokens),
                    "actualMin":  safe_number(row.actual_minutes),
                    "actualHHMM": minutes_to_hhmm(row.actual_minutes),
                    "targetMin":  safe_number(row.target_minutes),
                    "targetHHMM": minutes_to_hhmm(row.target_minutes),
                })
            modalities_dict[js_key] = points
        result.append({
            "key":        month_key(rep),
            "label":      month_label(rep),
            "modalities": modalities_dict,
        })
    return {"months": result}

def build_daily_status_summary(all_tokens: pd.DataFrame) -> dict:
    """
    Per-month daily token counts by status + modality breakdown.

    Each status bucket includes all variants of that base status —
    e.g. pendingTokens counts Pending, Pending - Done, Pending - NOT Done, etc.

    NOTE: completedTokens is intentionally exact 'Complete' only.
    eCompleteTokens includes all E-Complete variants.
    allCompletedTokens = completedTokens + eCompleteTokens.
    """
    def _sum_by_base(status_counts, base: str) -> int:
        """
        Sum all statuses whose base word matches, regardless of suffix variants.
        Safely handles E-Complete whose internal hyphen would otherwise be
        mistaken for a bucket separator.

        Works on both pd.Series (from value_counts()) and plain dict.
        """
        base_lower = base.strip().lower()
        total = 0
        for status, count in (status_counts.items() if hasattr(status_counts, 'items') else status_counts):
            s = str(status).strip()
            # Step 1: protect E-Complete's internal hyphen BEFORE normalizing separators
            s = re.sub(r'\bE\s*[-\.]\s*Complete\b', 'ECOMPLETE_PLACEHOLDER', s, flags=re.IGNORECASE)
            # Step 2: normalize all remaining ' - ' separators
            s = re.sub(r'\s*-\s*', ' - ', s)
            # Step 3: restore placeholder
            s = s.replace('ECOMPLETE_PLACEHOLDER', 'E-Complete')
            # Step 4: extract base (everything before first ' - ')
            base_part = s.split(' - ')[0].strip().lower()
            if base_part == base_lower:
                total += int(count)
        return total

    months  = _months_present(all_tokens)
    cleaned = all_tokens.dropna(subset=["date"]).copy()
    result  = []
    for rep in months:
        subset = cleaned[
            cleaned["date"].apply(
                lambda d: d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby(["date", "status"], dropna=False)
            .size()
            .reset_index(name="count")
            .sort_values(["date", "status"])
        )
        daily_counts = []
        for date_val, date_frame in grouped.groupby("date"):
            date_str     = date_val.isoformat() if isinstance(date_val, date) else str(date_val)
            status_counts: dict[str, int] = {}
            total_tokens = 0
            for row in date_frame.itertuples(index=False):
                status = row.status if row.status else "Unknown"
                count  = int(row.count)
                status_counts[status] = count
                total_tokens         += count

            # Exact match for Complete (Complete - Done goes to nonCompleteDone)
            completed_exact  = status_counts.get("Complete", 0)
            # All E-Complete variants
            ecomplete_all    = _sum_by_base(status_counts, "E-Complete")
            all_completed    = completed_exact 

            nc_done = sum(c for s, c in status_counts.items() if is_non_complete_done(s))
            nc      = sum(c for s, c in status_counts.items() if is_non_complete(s))

            completion_rate = (all_completed / total_tokens * 100) if total_tokens > 0 else 0.0

            daily_counts.append({
                "date":                  date_str,
                "totalTokens":           total_tokens,
                "statusCounts":          status_counts,
                "completedTokens":       completed_exact,
                "eCompleteTokens":       ecomplete_all,
                "allCompletedTokens":    all_completed,
                "nonCompleteDoneTokens": nc_done,
                "nonCompletedTokens":    nc,
                "pendingTokens":         _sum_by_base(status_counts, "Pending"),
                "servingTokens":         _sum_by_base(status_counts, "Serving"),
                "noShowTokens":          _sum_by_base(status_counts, "Noshow"),
                "standbyTokens":         _sum_by_base(status_counts, "Standby"),
                "completionRate":        safe_number(completion_rate, 1),
            })

        modality_grouped = (
            subset[subset["modality"] != "Unknown"]
            .groupby(["date", "modality", "status"], dropna=False)
            .size()
            .reset_index(name="count")
        )
        daily_modality: dict[str, dict] = {}
        for date_val, dmf in modality_grouped.groupby("date"):
            date_str = date_val.isoformat() if isinstance(date_val, date) else str(date_val)
            daily_modality[date_str] = {}
            for modality, mf in dmf.groupby("modality"):
                comp    = int(mf[mf["status"].apply(is_completed)]["count"].sum())
                nc_done = int(mf[mf["status"].apply(is_non_complete_done)]["count"].sum())
                nc      = int(mf[mf["status"].apply(is_non_complete)]["count"].sum())
                daily_modality[date_str][modality] = {
                    "completed":       comp,
                    "nonCompleteDone": nc_done,
                    "nonCompleted":    nc,
                }

        result.append({
            "key":           month_key(rep),
            "label":         month_label(rep),
            "daily":         daily_counts,
            "dailyModality": daily_modality,
        })
    return {"months": result}


def build_daily_process_breakdown(completed_tokens: pd.DataFrame) -> dict:
    """
    Per-month daily averages for service and billing stage components.

    FIX — Weighted averages: each point now carries a 'tokens' count so the
    frontend can compute a proper token-weighted mean across days, instead of
    an unweighted average-of-daily-averages (which gives equal weight to a
    2-token day and a 50-token day).

    FIX — Reconciliation: componentTotal is emitted for both service and
    billing breakdowns so the frontend can compare them against the computed
    TAT column and surface any Excel formula gaps.
    """
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        completed_tokens["date"].notna() &
        (completed_tokens["modality"] != "Unknown")
    ].copy()

    stage_columns = {
        "Ultrasound": ("us_stage_wait_minutes",  "us_stage_service_minutes",  "us_stage_hold_minutes"),
        "XRAY":       ("xr_stage_wait_minutes",  "xr_stage_service_minutes",  "xr_stage_hold_minutes"),
        "CT":         ("ct_stage_wait_minutes",  "ct_stage_service_minutes",  "ct_stage_hold_minutes"),
        "MRI":        ("mr_stage_wait_minutes",  "mr_stage_service_minutes",  "mr_stage_hold_minutes"),
    }
    billing_columns = {
        "Ultrasound": ("us_billing_wait_minutes", "us_billing_service_minutes", "us_billing_hold_minutes"),
        "XRAY":       ("xr_billing_wait_minutes", "xr_billing_service_minutes", "xr_billing_hold_minutes"),
        "CT":         ("ct_billing_wait_minutes", "ct_billing_service_minutes", "ct_billing_hold_minutes"),
        "MRI":        ("mr_billing_wait_minutes", "mr_billing_service_minutes", "mr_billing_hold_minutes"),
    }

    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        modalities = []
        for modality in sorted(subset["modality"].dropna().unique()):
            modality_code      = modality_code_from_label(modality)
            if modality_code is None:
                continue
            service_column     = SERVICE_COLUMNS_BY_MODALITY.get(modality_code)
            billing_tat_column = BILLING_COLUMNS_BY_MODALITY.get(modality_code)
            if not service_column or not billing_tat_column:
                continue
            stage_cols   = stage_columns.get(modality)
            billing_cols = billing_columns.get(modality)
            if not stage_cols or not billing_cols:
                continue
            mf = subset[subset["modality"] == modality].copy()
            grouped = (
                mf.groupby("date", dropna=False)
                .agg(
                    tokens=("token", "count"),
                    billing_minutes=(f"{billing_tat_column}_minutes", "mean"),
                    service_minutes=(f"{service_column}_minutes",     "mean"),
                    dispatch_minutes=("dispatch_tat_minutes",          "mean"),
                    service_wait=   (stage_cols[0],   "mean"),
                    service_service=(stage_cols[1],   "mean"),
                    service_hold=   (stage_cols[2],   "mean"),
                    billing_wait=   (billing_cols[0], "mean"),
                    billing_service=(billing_cols[1], "mean"),
                    billing_hold=   (billing_cols[2], "mean"),
                )
                .reset_index()
                .sort_values("date")
            )
            points = []
            for row in grouped.itertuples(index=False):
                svc_w = safe_number(row.service_wait)
                svc_s = safe_number(row.service_service)
                svc_h = safe_number(row.service_hold)
                bil_w = safe_number(row.billing_wait)
                bil_s = safe_number(row.billing_service)
                bil_h = safe_number(row.billing_hold)
                svc_total = safe_number(sum(v for v in [svc_w, svc_s, svc_h] if v is not None))
                bil_total = safe_number(sum(v for v in [bil_w, bil_s, bil_h] if v is not None))
                points.append({
                    "date":            row.date.isoformat() if isinstance(row.date, date) else str(row.date),
                    "tokens":          int(row.tokens),          # ← for weighted avg in JS
                    "billingMinutes":  safe_number(row.billing_minutes),
                    "serviceMinutes":  safe_number(row.service_minutes),
                    "dispatchMinutes": safe_number(row.dispatch_minutes),
                    "serviceBreakdown": {
                        "waitMinutes":    svc_w,
                        "serviceMinutes": svc_s,
                        "holdMinutes":    svc_h,
                        "componentTotal": svc_total,  # reconciliation vs serviceMinutes
                    },
                    "billingBreakdown": {
                        "waitMinutes":    bil_w,
                        "serviceMinutes": bil_s,
                        "holdMinutes":    bil_h,
                        "componentTotal": bil_total,  # reconciliation vs billingMinutes
                    },
                })
            modalities.append({"modality": modality, "points": points})
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
    print(f"  ✓  {path.relative_to(BASE_DIR)}")


def parse_iso_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            f"Invalid date '{value}'. Use YYYY-MM-DD format."
        ) from exc


def apply_date_filter(
    frame: pd.DataFrame,
    start_date: date | None,
    end_date: date | None,
) -> pd.DataFrame:
    filtered = frame
    if start_date is not None:
        filtered = filtered[filtered["date"] >= start_date]
    if end_date is not None:
        filtered = filtered[filtered["date"] <= end_date]
    return filtered


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export dashboard JSON files from TAT Excel sources."
    )
    parser.add_argument("--all-tokens",       dest="all_tokens_path",       default=str(ALL_TOKENS_PATH))
    parser.add_argument("--completed-tokens",  dest="completed_tokens_path",  default=str(COMPLETED_TOKENS_PATH))
    parser.add_argument("--output-dir",        dest="output_dir",             default=str(OUTPUT_DIR))
    parser.add_argument("--sheet-name",        dest="sheet_name",             default=DATA_SHEET)
    parser.add_argument("--start-date",        dest="start_date",             type=parse_iso_date, default=None)
    parser.add_argument("--end-date",          dest="end_date",               type=parse_iso_date, default=None)
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = build_args()
    if args.start_date and args.end_date and args.start_date > args.end_date:
        raise SystemExit("--start-date must be earlier than or equal to --end-date")

    global DATA_SHEET
    DATA_SHEET = args.sheet_name

    print("Reading Excel files…")
    all_tokens       = prepare_all_tokens(Path(args.all_tokens_path))
    completed_tokens = prepare_completed_tokens(Path(args.completed_tokens_path))

    if args.start_date or args.end_date:
        all_tokens       = apply_date_filter(all_tokens,       args.start_date, args.end_date)
        completed_tokens = apply_date_filter(completed_tokens, args.start_date, args.end_date)

    output_dir = Path(args.output_dir)
    print(f"\nWriting JSON to {output_dir.relative_to(BASE_DIR)}/")

    outputs = {
        "scorecards.json":              build_scorecards(all_tokens),
        "modality_status.json":         build_modality_status(all_tokens),
        "tat_vs_target.json":           build_tat_vs_target(completed_tokens),
        "tat_distribution.json":        build_tat_distribution(completed_tokens),
        "daily_trends.json":            build_daily_trends(completed_tokens),
        "daily_status_summary.json":    build_daily_status_summary(all_tokens),
        "daily_process_breakdown.json": build_daily_process_breakdown(completed_tokens),
    }

    for file_name, payload in outputs.items():
        write_json(output_dir / file_name, payload)

    if _UNKNOWN_MODALITY_LABELS:
        labels = ", ".join(sorted(_UNKNOWN_MODALITY_LABELS))
        print(
            f"\n  ⚠  Skipped unknown modality codes: {labels}\n"
            "     These are not in MODALITY_LABELS and have no TAT columns defined.\n"
            "     Add them to MODALITY_LABELS (and the corresponding column aliases)\n"
            "     if you want them to appear in the TAT distribution / breakdown charts."
        )

    print("\nDone. All JSON files updated.")


if __name__ == "__main__":
    main()